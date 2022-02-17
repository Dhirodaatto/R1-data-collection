import speedtest
# import matplotlib.pyplot as plt
import openpyxl as opx
from datetime import datetime
import os.path

def mbps(measure):
    return [i/1024/1024 for i in measure]

def get_reading():
    wifi_data = speedtest.Speedtest()
    wifi_data = mbps([wifi_data.download(), wifi_data.upload()])
    return wifi_data

def mean(test_list):
    mean = sum(test_list) / len(test_list)
    return mean

def stddev(test_list):
    variance = sum([((x - mean(test_list)) ** 2) for x in test_list]) / len(test_list)
    res = variance ** 0.5
    return res

more_readings = True

reading_set = 5

print("*** R1 Data Collection ***")
print("Warning! If you had paused data collection at some time, do enter a different name for the sheet")
print("so that your previous data does not get overriden !!")
print()

name = input("Please enter your Name here : ")
sheetname = input("Please enter the name of the sheet you wish to write to : ")
filepath = input("Please enter your filepath here : ")

if (os.path.exists(f'{filepath}/R1_reading_data_{name}.xlsx')):
    workbook = opx.load_workbook(f'{filepath}/R1_reading_data_{name}.xlsx')
else:
    workbook = opx.Workbook()

ws = workbook.create_sheet(sheetname)

ws.cell(row = 1, column=1).value = "Timestamp for reading"
ws.cell(row = 1, column=2).value = "Measured Download speed [Mbps]"
ws.cell(row = 1, column=3).value = "Measured upload speed [Mbps]"

ws.cell(row = 1, column=4).value = "Mean of Download set readings"
ws.cell(row = 1, column=5).value = "Mean of Upload set readings"
ws.cell(row = 1, column=6).value = "Standard deviation of Download set readings"
ws.cell(row = 1, column=7).value = "Standard deviation of Upload set readings"
ws.cell(row = 1, column=8).value = "Comments"

uc, dc = 5, 5

r = 1

dplot = list()
uplot = list()
xax = list()

while more_readings:
    
    rowset = r * reading_set + 1
    
    print()
    comments = input("Do you wish to enter any comments in the excel file for this set of readings ? ")
    print('Set Reading Started')
    
    ws.cell(row = rowset, column = 8).value = comments
    
    current_set_dspeed, current_set_uspeed = [], []
    
    for r_num in range(0, reading_set):
        dateobj = datetime.now()
        dspeed, uspeed = get_reading()
        ws.cell(row = rowset + r_num, column = 1).value = dateobj.date().strftime("%b %d %Y ") + dateobj.time().strftime("%H:%M:%S.%f")
        
        ws.cell(row = rowset + r_num, column = 2).value = dspeed
        ws.cell(row = rowset + r_num, column = 3).value = uspeed
        
        current_set_uspeed.append(uspeed)
        current_set_dspeed.append(dspeed)
        
        print(f'Set Reading number {r_num} done for your current reading -> {r}')
    
    ws.cell(row = rowset, column = 4).value = mean(current_set_dspeed)
    ws.cell(row = rowset, column = 5).value = mean(current_set_uspeed)
    ws.cell(row = rowset, column = 6).value = stddev(current_set_dspeed)
    ws.cell(row = rowset, column = 7).value = stddev(current_set_uspeed)
    
    print(f'Reading {r} done')
    test = input("Press Enter to continue and QQ to finish data collection  ")
    if (test == 'QQ'):
        break
    print()
    
    r = r + 1

# plt.plot([i for i in range(5)], dplot)
# plt.show()

workbook.save(filename = f'{filepath}/R1_reading_data_{name}.xlsx')    
